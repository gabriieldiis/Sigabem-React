import os
import datetime
from openpyxl import load_workbook

# Carregar o arquivo Excel
excel_file_path = r'C:\_projetos\python\Kalung\_rodrigo\htmls-loja-pj.xlsx'
workbook = load_workbook(excel_file_path)
sheet = workbook.active
date = datetime.datetime.now()

# Diretório onde os arquivos HTML serão salvos
output_directory = r'C:\_projetos\python\Kalung\_rodrigo\PJ'

# Loop pelas linhas do Excel
for row in sheet.iter_rows(min_row=1, values_only=True):
    content = row[0]  # Conteúdo da coluna A
    filename = row[1] + row[2]  # Nome do arquivo + extensão (.html)
    file_path = os.path.join(output_directory, filename)
    
    # Criar o arquivo HTML com o conteúdo da coluna A
    with open(file_path, 'w', encoding='utf-8') as file:
        file.write(content)
    
    print(f"Arquivo '{filename}' criado.")

print("Conversão concluída.")

print(date.strftime("%H")) 
print("--")
print(date.strftime("%M")) 